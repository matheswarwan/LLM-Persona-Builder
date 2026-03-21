"""
Extract persona data from JJ's LLM Personas xlsx into src/data/personas.json
Run: python3 extract_data.py
Requires: pip install openpyxl
"""

import json
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = "JJ's LLM Personas_v1.7.xlsx"
OUT_PATH = os.path.join("src", "data", "personas.json")

# Map Dashboard cell references → template tokens
DASHBOARD_TOKEN_MAP = {
    "O1": "{userName}",
    "O2": "{assistantName}",
    "B6": "{mood}",
    "B7": "{approach}",
    "B8": "{randomness}",
    "B27": "{randomFactsTopics}",
    "B28": "{storyAuthors}",
}


def formula_to_template(formula: str) -> str:
    """Replace Dashboard!$X$Y references with {token} placeholders."""
    if not formula:
        return formula

    def replace_ref(match):
        col = match.group(1)
        row = match.group(2)
        key = f"{col}{row}"
        return DASHBOARD_TOKEN_MAP.get(key, match.group(0))

    # Replace Dashboard!$COL$ROW patterns
    result = re.sub(
        r'Dashboard!\$([A-Z]+)\$(\d+)',
        replace_ref,
        formula
    )

    # Strip surrounding string concatenation artifacts: " & ... & "
    # The formula looks like: "prefix text" & Dashboard!$O$2 & " more text"
    # We want to convert it to: prefix text{token} more text
    # Evaluate the string parts by removing the Excel formula syntax
    result = re.sub(r'"\s*&\s*', '', result)  # "text" &  -> text
    result = re.sub(r'\s*&\s*"', '', result)  # & "text  -> text
    result = re.sub(r'^"', '', result)          # leading "
    result = re.sub(r'"$', '', result)          # trailing "
    # Clean up any remaining & operators between tokens
    result = re.sub(r'\s*&\s*', '', result)
    # Handle CHAR(10) → newline
    result = result.replace('CHAR(10)', '\n')

    return result


def cell_value(cell) -> str | None:
    """Get the display value of a cell (string or number as string)."""
    if cell.value is None:
        return None
    v = cell.value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, (int, float)):
        return str(v)
    return str(v)


def cell_template(cell) -> str | None:
    """
    Get the template string for a cell.
    If the cell has a formula, convert it to a template with {tokens}.
    Otherwise return the static value.
    """
    if cell.data_type == 'f' or (hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('=')):
        # Has formula - but openpyxl with data_only=False stores formula in value
        pass

    # We need two passes: one with data_only=False for formulas, one with data_only=True for values
    return None  # handled below


def extract_persona_rows(wb_formulas, wb_values):
    """Extract rows from Persona_Data sheet."""
    ws_f = wb_formulas["Persona_Data"]
    ws_v = wb_values["Persona_Data"]

    rows = []
    for row_idx in range(2, ws_f.max_row + 1):
        # Column mapping (1-indexed):
        # A=1: #/ID
        # C=3: Section code
        # D=4: Persona name
        # F=6: Section header text
        # G=7: Label
        # H=8: Value (may have formula)
        # J=10: Assembled prompt (formula)

        section_cell = ws_v.cell(row=row_idx, column=3)
        persona_cell = ws_v.cell(row=row_idx, column=4)
        header_cell = ws_v.cell(row=row_idx, column=6)
        label_cell = ws_v.cell(row=row_idx, column=7)
        value_cell_v = ws_v.cell(row=row_idx, column=8)
        value_cell_f = ws_f.cell(row=row_idx, column=8)

        section = cell_value(section_cell)
        if not section:
            continue

        persona = cell_value(persona_cell)
        # Normalize: blank persona = universal
        if not persona or persona.strip() == '':
            persona = None

        header = cell_value(header_cell)
        label = cell_value(label_cell)
        value_static = cell_value(value_cell_v)

        # Get formula for value cell
        formula_raw = value_cell_f.value
        value_template = None
        if isinstance(formula_raw, str) and formula_raw.startswith('='):
            formula_body = formula_raw[1:]  # strip leading =
            converted = formula_to_template(formula_body)
            # Only use template if it contains actual tokens
            if '{' in converted:
                value_template = converted
                # Also clean up extra quotes that may remain
                value_template = value_template.replace('""', '"')
            # If no tokens, use the static value

        # Skip rows with no useful content
        if not header and not label and not value_static and not value_template:
            continue

        rows.append({
            "section": section,
            "persona": persona,
            "sectionHeader": header,
            "label": label,
            "value": value_static,
            "valueTemplate": value_template,
        })

    return rows


def extract_lists(wb_values):
    """Extract dropdown options from Lists sheet."""
    ws = wb_values["Lists"]

    personas = []
    moods = []
    approaches = []
    sections = []

    # Row 1 is header; data starts at row 2
    # Column A: Unique Priorities (sections)
    # Column C: Unique Personas
    # Column E: Mood name
    # Column F: Mood description
    # Column H: Approach name
    # Column I: Approach description

    for row_idx in range(2, ws.max_row + 1):
        sec = cell_value(ws.cell(row=row_idx, column=1))
        persona = cell_value(ws.cell(row=row_idx, column=3))
        mood_name = cell_value(ws.cell(row=row_idx, column=5))
        mood_desc = cell_value(ws.cell(row=row_idx, column=6))
        approach_name = cell_value(ws.cell(row=row_idx, column=8))
        approach_desc = cell_value(ws.cell(row=row_idx, column=9))

        if sec and sec not in [s["code"] for s in sections]:
            sections.append({"code": sec})
        if persona and persona not in personas:
            personas.append(persona)
        if mood_name and mood_name not in [m["name"] for m in moods]:
            moods.append({"name": mood_name, "description": mood_desc or ""})
        if approach_name and approach_name not in [a["name"] for a in approaches]:
            approaches.append({"name": approach_name, "description": approach_desc or ""})

    return personas, moods, approaches, sections


def extract_bonus_prompts(wb_values):
    """Extract bonus prompt snippets from Bonus Prompts sheet."""
    ws = wb_values["Bonus Prompts"]
    bonus = []
    current_category = "General"

    for row_idx in range(1, ws.max_row + 1):
        val = cell_value(ws.cell(row=row_idx, column=3))
        if not val:
            continue

        # Category headers are short labels (no period, no quotes typically)
        # Heuristic: if the row has no other content and is short
        # Known categories from the data: "Constraint/Rule", "Format/Structure",
        # "Process/Methodology", "Tone/Persona", "Bonus Guidelines!"
        known_categories = [
            "Bonus Guidelines!", "Constraint/Rule", "Format/Structure",
            "Process/Methodology", "Tone/Persona"
        ]
        if val in known_categories:
            current_category = val
            continue
        # Skip short annotation rows like "JJ's fav:"
        if val.endswith(":") and len(val) < 30:
            continue

        bonus.append({"category": current_category, "text": val})

    return bonus


def extract_defaults(wb_values):
    """Extract default values from Dashboard sheet."""
    ws = wb_values["Dashboard"]
    defaults = {
        "userName": cell_value(ws["O1"]) or "JJ",
        "assistantName": cell_value(ws["O2"]) or "G",
        "persona": cell_value(ws["B5"]) or "# Salesforce Solution Architect",
        "mood": cell_value(ws["B6"]) or "Stoic",
        "approach": cell_value(ws["B7"]) or "Brief",
        "randomness": cell_value(ws["B8"]) or "3-8",
        "randomFactsTopics": cell_value(ws["B27"]) or "",
        "storyAuthors": cell_value(ws["B28"]) or "",
        "location": "Canada",
    }
    # Default section toggles (col A = include flag, col B = section code)
    section_defaults = {}
    for row_idx in range(12, 25):
        flag_cell = ws.cell(row=row_idx, column=1)
        code_cell = ws.cell(row=row_idx, column=2)
        flag = cell_value(flag_cell)
        code = cell_value(code_cell)
        if code:
            section_defaults[code] = (flag == "1" or flag == "TRUE" or flag is True)
    defaults["sectionDefaults"] = section_defaults
    return defaults


SECTION_DESCRIPTIONS = {
    "01-Persona": "Persona, Tone, Role, Primary / Secondary objectives, etc.",
    "02-All": "Task, content, constraints, behavior, formatting, final responses.",
    "04-Concise": "Turn on 'brief' mode — less noise. Turn off for more 'human' responses.",
    "05-Technical": "Assist in debugging; sorts response into sections, goes through each one by one.",
    "10-Facts": "Random fact that appears at the end of 1-n responses.",
    "11-A Story": "Random story moment that appears at the end of 1-n responses.",
    "12-Pro Tips": "Pro-tips on the current discussion.",
    "13-Architects' Note": "Architect's notes on the current discussion.",
    "20-Cmds_Sys": "System commands, ie !sync, !debug.",
    "21-Cmds-Docs": "Documentation commands, ie !import_PDF, !import_exam.",
    "22-Cmds-Misc": "Miscellaneous commands, ie !convert_table.",
    "25-Alter-ego": "Add an alter-ego B that gives brutally honest thoughts on your work.",
    "99-References": "Templates to provide examples for responses to imitate.",
}


def main():
    print(f"Loading {XLSX_PATH}...")
    wb_formulas = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    wb_values = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("Extracting persona rows...")
    rows = extract_persona_rows(wb_formulas, wb_values)
    print(f"  → {len(rows)} rows extracted")

    print("Extracting lists...")
    personas, moods, approaches, sections = extract_lists(wb_values)
    print(f"  → {len(personas)} personas, {len(moods)} moods, {len(approaches)} approaches, {len(sections)} sections")

    print("Extracting bonus prompts...")
    bonus_prompts = extract_bonus_prompts(wb_values)
    print(f"  → {len(bonus_prompts)} bonus prompts")

    print("Extracting defaults...")
    defaults = extract_defaults(wb_values)

    # Enrich sections with descriptions and defaults
    for s in sections:
        s["description"] = SECTION_DESCRIPTIONS.get(s["code"], s["code"])
        s["default"] = defaults["sectionDefaults"].get(s["code"], True)

    data = {
        "personas": personas,
        "moods": moods,
        "approaches": approaches,
        "sections": sections,
        "rows": rows,
        "defaults": {
            "userName": defaults["userName"],
            "assistantName": defaults["assistantName"],
            "location": defaults["location"],
            "persona": defaults["persona"],
            "mood": defaults["mood"],
            "approach": defaults["approach"],
            "randomness": defaults["randomness"],
            "randomFactsTopics": defaults["randomFactsTopics"],
            "storyAuthors": defaults["storyAuthors"],
        },
        "bonusPrompts": bonus_prompts,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\nDone! Output written to {OUT_PATH}")
    print(f"Summary:")
    print(f"  Personas:      {len(personas)}")
    print(f"  Moods:         {len(moods)}")
    print(f"  Approaches:    {len(approaches)}")
    print(f"  Sections:      {len(sections)}")
    print(f"  Prompt rows:   {len(rows)}")
    print(f"  Bonus prompts: {len(bonus_prompts)}")


if __name__ == "__main__":
    main()
